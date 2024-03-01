import os
import random
import numpy as np
import cv2
from tracker import Tracker
import cv2
from inference import get_roboflow_model
import time
import openpyxl
import math
import copy
from sklearn.cluster import KMeans
from sklearn import metrics
import matplotlib.pyplot as plt

colors = [[random.randint(0, 255) for _ in range(3)]
        for _ in range(10)]

FRAME_RATE = 30
BACK_FRAME_COUNT = int(FRAME_RATE * 0.75)


def add_car_record(track_id, cars, x1, y1, frame_count, min_cut=1):
    if track_id not in cars:
        cars[track_id] = {'init_pos':[(x1, y1, frame_count)], 'counter':1, 'final_pos':[]}
    else:
        cars[track_id]['counter'] += 1
        diff_x = abs(x1 - cars[track_id]['init_pos'][-1][0])
        diff_y = abs(y1 - cars[track_id]['init_pos'][-1][1])

        if len(cars[track_id]['init_pos']) < FRAME_RATE:
            if diff_x <= min_cut and diff_y <= min_cut:
                # the car has not moved
                cars[track_id]['init_pos'].pop(-1)
            cars[track_id]['init_pos'].append((x1, y1, frame_count))   
        
        cars[track_id]['final_pos'].append((x1, y1, frame_count))
        if len(cars[track_id]['final_pos']) > (FRAME_RATE + BACK_FRAME_COUNT):
            cars[track_id]['final_pos'].pop(0)

    return cars

def compute_angle(vector_x, vector_y):
    abs_x, abs_y = abs(vector_x), abs(vector_y)
    if abs_x != 0:
        angle = (np.arctan(abs_y/abs_x)/np.pi) * 180
    else:
        if abs_y == 0:
            raise ValueError(f'Expected a non-zero vector, but got <{abs_x}, {abs_y}>')
        return 90 if vector_y > 0 else 270
    if vector_x > 0 and vector_y > 0: # first quadrant
        pass
    elif vector_x < 0 and vector_y > 0: # second quadrant
        angle = 180 - angle
    elif vector_x < 0 and vector_y < 0: # third quadrant
        angle = 180 + angle
    elif vector_x > 0 and vector_y < 0: # fourth quadrant
        angle = 360 - angle
    elif vector_x < 0 and vector_y == 0:
        angle = 180
    elif vector_x > 0 and vector_y == 0:
        angle = 0
    return angle

def compute_car_vectors(cars, frame_width):
    """
    Computes the entry and exit vectors of each car in the cars dictionary
    """
    min_cut = 1 * frame_width/ 100
    new_cars = {}
    for car_id in cars:
        try:
            entry_vector_x = cars[car_id]['init_pos'][-1][0] - cars[car_id]['init_pos'][0][0]
            entry_vector_y = cars[car_id]['init_pos'][-1][1] - cars[car_id]['init_pos'][0][1]
            exit_vector_x = cars[car_id]['final_pos'][-BACK_FRAME_COUNT][0] - cars[car_id]['final_pos'][0][0]
            exit_vector_y = cars[car_id]['final_pos'][-BACK_FRAME_COUNT][1] - cars[car_id]['final_pos'][0][1]
            # change_vector_x = cars[car_id]['final_pos'][-BACK_FRAME_COUNT][0] - cars[car_id]['init_pos'][0]
            # change_vector_y = cars[car_id]['final_pos'][-BACK_FRAME_COUNT][1] - cars[car_id]['init_pos'][1]
            # distance_moved = np.sqrt(change_vector_x ** 2 + change_vector_y ** 2)
            if not ((abs(entry_vector_x) <= min_cut and abs(entry_vector_y) <=min_cut) \
                    or (abs(exit_vector_x) <= min_cut and abs(exit_vector_y) <= min_cut)):
                entry_angle = compute_angle(entry_vector_x, entry_vector_y)
                exit_angle = compute_angle(exit_vector_x, exit_vector_y)
                cars[car_id]['entry'] = round(entry_angle, 2)
                cars[car_id]['exit'] = round(exit_angle, 2)
                new_cars[car_id] = cars[car_id]
        except IndexError:
            # print('Index Errorr')
            pass
        except KeyError:
            # print('Key Error')
            pass
    # print(new_cars)
    return new_cars

def generate_stat_table(cars, filename):
    """
    Counts the number of cars that entered through every exit
    Counts the number of cars that exited via every exit
    Counts the number of cars that exited one particular exit z, coming from a particular exit y
    Writes all the data into an excel sheet

    Going counter-clockwise, we have the following bearings

    East Exit: 0 < angle <= 22.5 OR 337.5 < angle <= 360
    South East Exit:  22.5 < angle <= 67.5
    South Exit: 67.5 < angle <= 112.5
    South West Exit: 112.5 < angle <= 157.5

    West Exit: 157.5 < angle <= 202.5
    North West Exit: 202.5 < angle <= 247.5
    North Exit: 247.5 < angle <= 292.5
    North East Exit: 292.5 < angle <= 337.5
    """
    filename = filename + '_cars_id_directions' + '.xlsx'
    try:
        wb = openpyxl.load_workbook(
            filename
        )  # Open the spreadsheet for the genetic algorithm.
    except:
        wb = openpyxl.Workbook()
    
    bearings_count = {
        'West':[0, 0], 'South West':[0, 0], 'South':[0, 0], 'South East':[0, 0],
        'East':[0, 0], 'North East':[0, 0], 'North':[0, 0], 'North West':[0, 0]
    }

    combinations = { }

    for bearing in bearings_count:
        for bearing_2 in bearings_count:
            combinations[f'{bearing} - {bearing_2}'] = 0    
    row_counter = 1
    # Counting entrances and exits:
    for car_id in cars:
        row_counter += 1
        entry_angle = cars[car_id]['entry']
        exit_angle = cars[car_id]['exit']
        entry_b, exit_b = get_bearing(entry_angle, 'entry').title(), get_bearing(exit_angle).title()
        bearings_count[entry_b][0] += 1
        bearings_count[exit_b][1] += 1
        cmbn = f'{entry_b} - {exit_b}'
        combinations[cmbn] += 1
        wb["Sheet"].cell(row_counter, 1).value = car_id
        wb["Sheet"].cell(row_counter, 2).value = cmbn
        print(f'car with id {car_id} went {cmbn}')
    wb.save(filename)
    return [bearings_count, combinations]

def write_to_excel(bearings_count, combinations, filename):
    original_filename = filename
    filename = filename + '.xlsx'
    try:
        wb = openpyxl.load_workbook(
            filename
        )  # Open the spreadsheet for the genetic algorithm.
    except:
        wb = openpyxl.Workbook()
    
    bear_idx = {'West':0, 'North West':1, 'North':2, 'North East':3, 'East':4, 'South East': 5, 'South': 6, 'South West': 7}
    wb["Sheet"].cell(11, 1).value = "Bearing"
    wb["Sheet"].cell(11, 2).value = "Entry Counts"
    wb["Sheet"].cell(11, 3).value = "Exit Counts"
    for combi, count in combinations.items():
        row_name, col_name = combi.split(' - ')
        row_idx, col_idx = bear_idx[row_name], bear_idx[col_name]
        wb["Sheet"].cell(2 + row_idx, 1).value = row_name.title()
        wb["Sheet"].cell(1, 2 + col_idx).value = col_name.title()
        wb["Sheet"].cell(2 + row_idx, 2 + col_idx).value = count
    
    row_count = 11
    for bearing, counts_ in bearings_count.items():
        row_count += 1
        wb["Sheet"].cell(row_count, 1).value = bearing
        wb["Sheet"].cell(row_count, 2).value = counts_[0] # entry counts
        wb["Sheet"].cell(row_count, 3).value = counts_[1] # exit counts
    wb.save(filename)
    return original_filename

def encode(*args):
    assert len(args) == 4, f"Expected 4 arguments, got {len(args)}"
    mediate = []
    for arg in args:
        mediate.append(''.join(list(map(lambda x: x[0], arg.split(' ')))))
    return '-'.join(mediate)

def decode(code):
    mediate = code.split('-')
    assert len(mediate) == 4, f"Expected mediate to have 4 values but got {len(mediate)}"
    maping = {'N':'North', 'E':'East', 'S':'South', 'W':'West', \
                'NE':'North East', 'NW':'North West', 
                'SE':'South East', 'SW':'South West', 'C':'Center'}
    mediate_2 = list(map(lambda x: maping[x], mediate))
    return (mediate_2[0] + '-' + mediate_2[1], mediate_2[2] + '-' + mediate_2[3])

def write_cars_info_to_excel(cars, filename, img_width, img_height):
    filename = filename + "_track_info.xlsx"
    try:
        wb = openpyxl.load_workbook(
            filename
        )  # Open the spreadsheet for the genetic algorithm.
    except:
        wb = openpyxl.Workbook()
    
    row_count = 1
    for car_id in cars:
        row_count += 1
        final_x = cars[car_id]['final_pos'][-BACK_FRAME_COUNT][0]
        final_y = cars[car_id]['final_pos'][-BACK_FRAME_COUNT][1]
        init_x = cars[car_id]['init_pos'][0][0]
        init_y = cars[car_id]['init_pos'][0][1]
        entry_region = determine_region(img_width, img_height,\
            init_x, init_y)
        exit_region = determine_region(img_width, img_height,\
            final_x, final_y)
        entry = str(cars[car_id]['entry'])
        exit_ = str(cars[car_id]['exit'])
        entry_b = get_bearing(entry)
        exit_b = get_bearing(exit_)
        wb["Sheet"].cell(row_count, 1).value = car_id
        wb["Sheet"].cell(row_count, 2).value = str(cars[car_id]['init_pos'])
        wb["Sheet"].cell(row_count, 3).value = str(cars[car_id]['final_pos'])
        wb["Sheet"].cell(row_count, 4).value = entry
        wb["Sheet"].cell(row_count, 5).value = exit_
        wb["Sheet"].cell(row_count, 6).value = entry_region
        wb["Sheet"].cell(row_count, 7).value = exit_region
        wb["Sheet"].cell(row_count, 8).value = entry_b
        wb["Sheet"].cell(row_count, 9).value = exit_b
        wb["Sheet"].cell(row_count, 10).value = np.sqrt((final_x - init_x)**2 + (final_y - init_y)**2)
        wb["Sheet"].cell(row_count, 11).value = encode(entry_b, entry_region, exit_b, exit_region)
    wb["Sheet"].cell(1, 1).value = "car_id"
    wb["Sheet"].cell(1, 2).value = "init_pos"
    wb["Sheet"].cell(1, 3).value = "final_pos"
    wb["Sheet"].cell(1, 4).value = "entry"
    wb["Sheet"].cell(1, 5).value = "exit"
    wb["Sheet"].cell(1, 6).value = "entry region"
    wb["Sheet"].cell(1, 7).value = "exit region"
    wb["Sheet"].cell(1, 8).value = "entry_b"
    wb["Sheet"].cell(1, 9).value = "exit_b"
    wb["Sheet"].cell(1, 10).value = "dist covered"
    wb["Sheet"].cell(1, 11).value = "encoded_path"
    wb.save(filename)
    return {}

def read_cars_info_from_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print("File not found.")
        return {}

    sheet = wb.active
    cars = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        car_id = row[0]
        init_pos = eval(row[1])
        final_pos = eval(row[2])
        entry = eval(row[3])
        car_exit = eval(row[4])
        cars[car_id] = {'init_pos': init_pos, 'final_pos': final_pos, 'entry':entry,'exit':car_exit}

    return cars

def get_bearing(angle, direction="exit"):
    def reverse(direct):
        _dict = {'South':'North', 'North':'South', 'West':'East', 'East':'West'}
        return _dict[direct]
    dir_ = ""
    angle = float(angle)
    if (0 <= angle <= 22.5) or (337.5 < angle <= 360):
        dir_ = "East"
    elif (22.5 < angle <= 67.5):
        dir_ = "South East"
    elif (67.5 < angle <= 112.5):
        dir_ = "South"
    elif (112.5 < angle <= 157.5):
        dir_ = "South West"
    elif (157.5 < angle < 202.5):
        dir_ = "West"
    elif (202.5 < angle <= 247.5):
        dir_ = "North West"
    elif (247.5 < angle < 292.5):
        dir_ = "North"
    elif (292.5 < angle <= 337.5):
        dir_ = "North East"
    else:
        raise ValueError(f"Angle out of range: expected (0 < angle < 360) but got {angle}")
    """
    if direction == "exit":
        return dir_
    else:
        dir_ = dir_.split()
        result = list(map(lambda x:reverse(x), dir_))
        return " ".join(result)
    
    """
    return dir_

def draw_circle(frame, center, radius, color):
    # Draw circle on the frame
    cv2.circle(frame, center, radius, color, thickness=4)
    return frame

def label_car_on_frame(frame, track_id, x1, y1, x2, y2):
    cv2.rectangle(frame, (x1, y1), (x2, y2), (colors[track_id % len(colors)]), 3)
    label = track_id
    #Create a rectangle above the detected object and add label and confidence score
    t_size=cv2.getTextSize(str(label), cv2.FONT_HERSHEY_SIMPLEX, fontScale=1/2, thickness=1)[0]
    c2=x1+t_size[0], y1-t_size[1]-3
    cv2.rectangle(frame, (x1, y1), c2, color=(85, 45, 255), thickness=-1, lineType=cv2.LINE_AA)
    cv2.putText(frame, str(label), (x1, y1-2), 0, 1/2, [255, 255, 255], thickness=1, lineType=cv2.LINE_AA)
    return frame

def determine_region(img_width, img_height, x, y):
    regions_x = {0:'West', 1:'', 2:'East'}
    regions_y = {0:'North', 1:'', 2:'South'}
    x_index = math.floor((x/img_width) * (100/34))
    y_index = math.floor((y/img_height) * (100/34))
    
    if x_index > 2 or y_index > 2 or x_index < 0 or y_index < 0:
        print(f'Unexpected index, (x_index, y_index): \
              ({x_index}, {y_index}), img_size: ({img_width}, {img_height}), coordinates: ({x, y})') 
        if x_index > 2:
            x_index = 2 
        elif x_index < 0:
            x_index = 0
        if y_index > 2:
            y_index = 2 
        elif y_index < 0:
            y_index = 0 
    coord = regions_y[y_index] + " " + regions_x[x_index]
    coord = coord.strip()
    return coord if coord else "Center"

def process_video(folder_names):

    models = {'0':'center_of_intersection/5', '1':'outside_the_intersection/3', '2':'vaid-jy5xo/1'}
    model = get_roboflow_model(model_id=models['2'], api_key="UWNPpMrGPun13aPF9MLJ")

    for n in folder_names:
        st = time.perf_counter()
        filename = n
        extension = 'mp4'
        base_path = os.path.join('.', 'data', filename) 
        path_counter = 1
        specific_filename = os.path.join(base_path, f'{filename}_{path_counter}')
        video_path = os.path.join(base_path, f'{filename}_{path_counter}.{extension}')

        while os.path.exists(video_path): 
            video_out_path = os.path.join(base_path, f'{filename}_{path_counter}_out.{extension}')
            cap = cv2.VideoCapture(video_path)
            nframes = cap.get(cv2.CAP_PROP_FRAME_COUNT)

            """
            cars = {
            '0':{'init_pos':(x1, y1, frame_number_1),'counter':0, 
            'init_pos_2':(x2, y2, frame_number_2), 
            'final_pos':[(xn, yn, frame_number_n), ..., (xn+30, yn+30, frame_number_n+30)]}
            }
            """
            cars = {

            }


            print('there are ', nframes, 'frames')
            ret, frame = cap.read()
            first_frame = copy.deepcopy(frame)
            height, width = frame.shape[:2]
            cap_out = cv2.VideoWriter(video_out_path, cv2.VideoWriter_fourcc(*'MP4V'), cap.get(cv2.CAP_PROP_FPS),
                                    (frame.shape[1], frame.shape[0]))

            tracker = Tracker()
            detection_threshold = 0.3
            counter = 0
            frame_skip = 5
            while ret:
                result = model.infer(frame)
                detections = []
                for r in result[0].predictions:
                    score, class_id =  r.confidence, r.class_id
                    x1 =r.x -r.width / 2
                    x2 =r.x +r.width / 2
                    y1 =r.y -r.height / 2
                    y2 =r.y +r.height / 2
                    x1 = int(x1)
                    x2 = int(x2)
                    y1 = int(y1)
                    y2 = int(y2)
                    class_id = int(class_id)
                    if score > detection_threshold:
                        detections.append([x1, y1, x2, y2, score])

                tracker.update(frame, detections)

                for track in tracker.tracks:
                    bbox = track.bbox
                    x1, y1, x2, y2 = bbox
                    x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                    track_id = track.track_id
                    cars = add_car_record(track_id, cars, x1, y1, counter)  
                    frame = label_car_on_frame(frame, track_id, x1, y1, x2, y2)
                cap_out.write(frame)
                ret, frame = cap.read()
                counter += 1
            try:
                cap.release()
                cap_out.release()
                cv2.destroyAllWindows()
            except:
                print('encountered error')
            et = time.perf_counter()
            total_time_mins = (et - st)/60
            cars = compute_car_vectors(cars, width)
            nothing = write_cars_info_to_excel(cars=cars, filename=specific_filename, img_width=width, img_height=height)
            original_filename = write_to_excel(*generate_stat_table(cars, specific_filename), specific_filename)
            fr = vector_analysis(specific_filename, width, height)
            l = save_final_results(fr, specific_filename, first_frame, total_time_mins)
            
            print(f'Took {total_time_mins}mins')
            
            path_counter += 1  
            video_path = os.path.join(base_path, f'{filename}_{path_counter}.{extension}') # has to be after video_path
            specific_filename = os.path.join(base_path, f'{filename}_{path_counter}')

def is_border_zone(img_width, img_height, x, y):
    zone = 0.2
    percent_x = x/img_width
    percent_y = y/img_height
    valid_x, valid_y = False, False
    if percent_x <= zone or percent_x >= 1 - zone:
        valid_x = True
    if percent_y <= zone  or percent_y >= 1 - zone:
        valid_y = True
    return valid_x or valid_y

def vector_analysis(filename, img_width, img_height):

    def is_same(code):
        mediate = code.split('-')
        return (mediate[0] == mediate[2]) and (mediate[1] == mediate[3])
    
    filename = filename + "_track_info.xlsx"
    
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print("File not found.")
        return {}

    results = {}
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # print(row)
        car_id = row[0]
        init_x, init_y, __ = eval(row[1])[0]
        final_x, final_y, __ = eval(row[2])[0]
        if is_border_zone(img_width, img_height, init_x, init_y):
            dist = float(row[9])
            path_code = row[10]
            if not is_same(path_code):
                if path_code in results:
                    results[path_code][0].append(dist)
                    results[path_code][1].append(car_id)
                    results[path_code][2].append([(init_x, init_y), (final_x, final_y)])
                else:
                    results[path_code] = [[dist], [car_id], [[(init_x, init_y), (final_x, final_y)]]]
    final_results = {}
    for code, array in results.items():
        # print(array)
        std  = np.std(array[0])
        mean = np.mean(array[0])
        new_dist, ids, inits = [], [], []
        for index, distance in enumerate(array[0]):
            if std != 0 and ((distance - mean)/std) <= -2:
                pass
            else:
                new_dist.append(distance)
                ids.append(array[1][index])
                inits.append(array[2][index])
        final_results[code] = [new_dist, ids, inits]
    
    return final_results

def get_circles(x_y_positions):
    result = []
    for j in range(2):
        x_i_positions = [x_y_positions[i][j][0] for i in range(len(x_y_positions))]
        y_i_positions = [x_y_positions[i][j][1] for i in range(len(x_y_positions))]

        x_i_mean, y_i_mean = np.mean(x_i_positions), np.mean(y_i_positions)
        x_i_var, y_i_var = np.var(x_i_positions), np.var(y_i_positions)
        
        result.append((x_i_mean, y_i_mean, 3 * np.sqrt(x_i_var + y_i_var)))

    return result # (center_x, center_y, radius)


def draw_regions(code_circles, first_frame, filename):
    counter = -1
    for code, circles in code_circles.items():
        counter += 1
        for circle in circles:
            x, y, radius = circle
            center_coord = (int(x), int(y))
            first_frame = draw_circle(first_frame, center_coord, int(radius), colors[counter % 10])
    cv2.imwrite(f"{filename}_circles.jpg", first_frame)
    
    return None

def save_final_results(final_results, filename, first_frame, total_time, base=''):
    """
    """
    save_filename = os.path.join(base, filename + '_final_results.xlsx')
    try:
        wb = openpyxl.load_workbook(
            filename
        )  # Open the spreadsheet for the genetic algorithm.
    except:
        wb = openpyxl.Workbook()
    inner_row_counter = 1
    outer_row_counter = 1
    entrances = set()
    exits = set()
    code_count = {}
    code_circles = {}
    for code, array in final_results.items():
        outer_row_counter += 1
        for id, dist in zip(array[1], array[0]):
            inner_row_counter += 1
            wb["Sheet"].cell(inner_row_counter, 1).value = id
            wb["Sheet"].cell(inner_row_counter, 2).value = dist
            wb["Sheet"].cell(inner_row_counter, 3).value = code
        wb["Sheet"].cell(outer_row_counter, 6).value = code
        wb["Sheet"].cell(outer_row_counter, 7).value = len(array[0])
        code_circles[code] = get_circles(array[2])
        code_count[code] = len(array[0])
        et, ex = decode(code)
        entrances.add(et)
        exits.add(ex)

    foo = draw_regions(code_circles, first_frame, filename)
    exits = list(exits)
    entrances = list(entrances)
    for idx, ex in enumerate(exits):
        wb["Sheet"].cell(1, 10 + idx).value = ex
    for idy, et in enumerate(entrances):
        wb["Sheet"].cell(2 + idy, 10 - 1).value = et
    
    for idy, et in enumerate(entrances):
        for idx, ex in enumerate(exits):
            m = et.split('-') + ex.split('-')
            wb["Sheet"].cell(2 + idy, 10 + idx).value = code_count.get(encode(*m), 0)
    
    wb["Sheet"].cell(1, 1).value = "car ids"
    wb["Sheet"].cell(1, 2).value = "distances"
    wb["Sheet"].cell(1, 3).value = "car path code"
    wb["Sheet"].cell(1, 6).value = "Path Code"
    wb["Sheet"].cell(1, 7).value = "Count"
    wb["Sheet"].cell(20, 15).value = "Total time"
    wb["Sheet"].cell(21, 15).value = total_time
    wb.save(save_filename)
    return None

def get_frame_height(filename):
    extension = 'mp4'
    video_path = os.path.join('.', 'data', f'{filename}.{extension}')
    # video_out_path = os.path.join('.', f'main_{filename}_out.{extension}')
    cap = cv2.VideoCapture(video_path)   
    ret, frame = cap.read()
    width, height = frame.shape[:2]
    return height


def post_analysis(save_filename, excel_filename):
    filename = save_filename
    height = get_frame_height(filename=filename)
    cars = read_cars_info_from_excel(excel_filename)
    # cars = compute_car_vectors(cars, height) vectors should already be computed by now
    write_to_excel(*generate_stat_table(cars, filename), filename)



def get_cluster_n(inertias, ns):
    """
    Takes inertia values and ns [cluster numbers] and returns an estimate of 
    the best cluster number
    """
    assert len(inertias) == len(ns), "Inertias and number of n must be of same lengths"
    diffs = []
    for i in range(1, len(inertias)):
        diffs.append(inertias[i-1] - inertias[i])
    
    mean, std = np.mean(diffs), np.std(diffs)
    ids_values = []
    for idx, diff in enumerate(diffs):
        if (diff - mean) >= 0.2 * std:
            ids_values.append([idx, diff])
    # print(diffs, ids_values)
    min_idx = None
    min_val = None
    for id_value in ids_values:
        if min_val is None or id_value[1] < min_val: # must be strictly less than here
            min_val = id_value[1]
            min_idx = id_value[0]
    # Plot the elbow curve
    """
    plt.plot(ns, inertias, marker='o')
    plt.xlabel('Number of Clusters')
    plt.ylabel('Inertia')
    plt.title('Elbow Method')
    plt.show()
    """
    print(ns[min_idx + 1])
    return ns[min_idx + 1]

def compute_inertias(data, n_max=50):
    data = np.array(data)
    # Initialize lists to store inertia values for different numbers of clusters
    inertia_values = []
    # Range of clusters to try
    num_clusters_range = range(1, min(n_max + 1, len(data)))
    # print(num_clusters_range)

    # Calculate inertia (within-cluster sum of squares) for each number of clusters
    for num_clusters in num_clusters_range:
        kmeans = KMeans(n_clusters=num_clusters)
        kmeans.fit(data)
        inertia_values.append(kmeans.inertia_)
    return [inertia_values, num_clusters_range]

def classify_cluster(data):
    num_clusters = get_cluster_n(*compute_inertias(data))
    # num_clusters = 4
    X = np.array(data)
    # Initialntize the KMeans model
    kmeans = KMeans(n_clusters=min(num_clusters, len(data)))

    # Fit the model to the data
    kmeans.fit(X)

    # Get the cluster labels for each data point
    labels = kmeans.labels_

    # Get the cluster centers
    centers = kmeans.cluster_centers_
    
    clusters_data = {}
    # Calculate standard deviation for each cluster
    for i in range(num_clusters):
        clusters_data[i] = {'datapoints': X[labels == i], 'center': centers[i]}  # Filter data points belonging to cluster i
    
    for cluster_id in clusters_data:
        dp = clusters_data[cluster_id]['datapoints']
        x_points = dp[:, 0]
        y_points = dp[:, 1]
        radius = np.sqrt(np.var(x_points) + np.var(y_points))
        clusters_data[cluster_id]['radius'] = radius

    return clusters_data
    
def draw_regions_cluster(filename):

    extension = 'mp4'
    base = os.path.join('.', 'data', filename)
    video_path = os.path.join('.', 'data', filename, f'{filename}_1.{extension}')
    print(video_path)
    cap = cv2.VideoCapture(video_path)

    # Get the total number of frames in the video
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    # Set the frame position to the midpoint
    midpoint_frame = total_frames // 2
    cap.set(cv2.CAP_PROP_POS_FRAMES, midpoint_frame)
    # Read the frame at the midpoint
    ret, frame = cap.read()
    height, width = frame.shape[:2]

    counter = 0
    entry_data = []
    exit_data = []
    while True:
        counter += 1
        file = os.path.join(base, filename + '_' + str(counter) + "_track_info.xlsx")
        try:
            wb = openpyxl.load_workbook(file)
        except FileNotFoundError:
            break

        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                r1, r2 = eval(row[1])[3], eval(row[2])[0]
                ep = row[10] # encoded path
                if not 'C' in ep and is_border_zone(width, height, r1[0], r1[1]):
                    entry_data.append([r1[0], r1[1]])
                    exit_data.append([r2[0], r2[1]])
            except Exception as e:
                print(e)
    
    if not entry_data or not exit_data:
        raise NotADirectoryError('Could not find the files')
    
    entry_clusters = classify_cluster(entry_data)
    exit_clusters = classify_cluster(exit_data)
    # print(entry_clusters)
    #print(exit_clusters)
    clusters = [entry_clusters, exit_clusters]


    colors = [(100, 255, 100), (255, 50, 50)]
    for cluster, color in zip(clusters, colors):
        for cluster_id in cluster:
            center_coordinates = (int(cluster[cluster_id]["center"][0]), \
                                  int(cluster[cluster_id]["center"][1]))
            print(center_coordinates)
            radius = int(cluster[cluster_id]["radius"])
            print(cluster_id)
            frame = draw_circle_with_text(frame, center_coordinates, radius, str(cluster_id), color=color)

    save_path = os.path.join(base, f"{filename}_cluster_circles.jpg")
    
    cv2.imwrite(save_path, frame)

    return clusters


def write_final_results(filename, clusters):
        base = os.path.join('.', 'data', filename) 
        counter = 0
        entry_cluster, exit_cluster = clusters[0], clusters[1]
        c1_len, c2_len = len(entry_cluster), len(exit_cluster)
        count_array = [[0 for j in range(c2_len)] for _ in range(c1_len)]


        while True:
            counter += 1
            file = os.path.join(base, filename + '_' + str(counter) + "_track_info.xlsx")
            try:
                wb = openpyxl.load_workbook(file)
            except FileNotFoundError:
                    break   
            sheet = wb.active
            row_counter = 1
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_counter += 1
                try:
                    r1, r2 = eval(row[1])[3], eval(row[2])[0]
                    car_id = int(row[0])
                    # print(r1, r2)
                    ep = row[10] # encoded path
                    car_entry_id, car_exit_id = None, None
                    for eid in entry_cluster:
                        entry_dist = (r1[0] - entry_cluster[eid]["center"][0])**2 + \
                                (r1[1] - entry_cluster[eid]["center"][1])**2 
                        # if car_id == 40: print(r1, entry_dist, 2.25 * entry_cluster[eid]["radius"]**2)
                        if entry_dist <= 2.25 * entry_cluster[eid]["radius"]**2:
                            car_entry_id = eid
                            for exid in exit_cluster:
                                exit_dist = (r2[0] - exit_cluster[exid]["center"][0])**2 + \
                                        (r2[1] - exit_cluster[exid]["center"][1])**2
                                # if car_id == 40: print(r2, exit_dist, 2.25 * exit_cluster[exid]["radius"]**2)
                                if exit_dist <= 2.25 * exit_cluster[exid]["radius"]**2:
                                    car_exit_id = exid
                                    count_array[car_entry_id][car_exit_id] += 1
                                    break
                            break
                    sheet.cell(row_counter, 12).value = str(car_entry_id) + '-' + str(car_exit_id)
                    # print(entry_dist, exit_dist)
                except Exception as e:
                    print(e)
                
            
            wb.save(file)
        new_wb = openpyxl.Workbook()

        for j in range(c2_len):
            new_wb["Sheet"].cell(1, j + 2).value = j
        for i in range(c1_len):
            new_wb["Sheet"].cell(i + 2, 1).value = i
        
        for idx, sub_array in enumerate(count_array):
            for idy, val in enumerate(sub_array):
                new_wb["Sheet"].cell(idx + 2, idy + 2).value = val
        new_wb.save(os.path.join(base, filename + "_table.xlsx"))
                
        return None

# Function to draw circle and write text
def draw_circle_with_text(frame, center_coordinates, radius, text, color=(0, 255, 0)):
    # Draw circle
    cv2.circle(frame, center_coordinates, radius, color , 2)
    
    # Write text inside circle
    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.5
    font_thickness = 2
    text_size, _ = cv2.getTextSize(text, font, font_scale, font_thickness)
    text_x = center_coordinates[0] - text_size[0] // 2
    text_y = center_coordinates[1] + text_size[1] // 2
    cv2.putText(frame, text, (text_x, text_y), font, font_scale, color, font_thickness)
    # cv2.imwrite(f"{filename}_cluster_circles.jpg", frame)  

    return frame

def print_frames():
    vids = ['roundabout_test','1', 'Low Cost Implanted Compact Roundabout – Operation – Drone Footage', 'Intersection Traffic Study _ Florida Aerial Mapping Services']
    # vids = ['valpo_test_roundabout']
    for v in vids:
        draw_regions_cluster(v)

if __name__=='__main__':
    folder_names = ['1']
    # process_video(folder_names)
    for fn in folder_names:
        clusters = draw_regions_cluster(fn)
        empty = write_final_results(fn, clusters)
